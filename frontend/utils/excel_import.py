"""
Helpers for reading Excel files used by import flows.
"""
from __future__ import annotations

import re
from typing import Any


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = (
        text.replace("ã", "a")
        .replace("á", "a")
        .replace("à", "a")
        .replace("â", "a")
        .replace("é", "e")
        .replace("ê", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ô", "o")
        .replace("õ", "o")
        .replace("ú", "u")
        .replace("ç", "c")
    )
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def normalize_identifier(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def normalize_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def pick_first_value(row: dict[str, Any], candidates: list[str], default: Any = None):
    for key in candidates:
        if key in row and row[key] not in (None, ""):
            return row[key]
    return default


def read_excel_rows(file_path: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("Dependencia ausente para importacao Excel: instale openpyxl.") from exc

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        wb.close()
        return []

    headers = [normalize_header(h) for h in (headers_raw or [])]
    result: list[dict[str, Any]] = []
    for values in rows_iter:
        if not values:
            continue
        row: dict[str, Any] = {}
        has_data = False
        for idx, value in enumerate(values):
            key = headers[idx] if idx < len(headers) else f"coluna_{idx+1}"
            if not key:
                key = f"coluna_{idx+1}"
            if value not in (None, ""):
                has_data = True
            row[key] = value
        if has_data:
            result.append(row)
    wb.close()
    return result

